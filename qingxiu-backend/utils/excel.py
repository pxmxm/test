import os

import openpyxl

from backend.settings import MEDIA_ROOT, WEB_HOST_MEDIA_URL

a = [{
    "annualPlan": "2021",
    "projectBatch": "第N批次",
    "planCategory": "科技重大专项",
    "projectName": "人工智能和电子信息产业1",
    "subjectName": "人工智能和电子信息申报课题1",
    "unitName": "青秀区科技计划管理系统测试单位、一起申报有限公司",
    "head": "钱悦",
    "overallGoal": "课题总体目标是完成人工智能和电子信息工程涉及",
    "startStopYear": "2021.06-2022.12",
    "assessmentIndicators": "考核指标1：完成第一阶段目标\n考核指标2：完成第二阶段目标\n考核指标x：完成第三阶段目标",
    "scienceFunding": 40000000,
    "unitSelfRaised": 0,
    "subjectState": "专家评审",
    "a": {
        "subjectScore": "-",
        "proposal": "-",
        "proposalFunding": "-"
    },
    "b": {
        "scienceProposal": "-",
        "scienceFunding": 0
    },
    "c": {
        "executionTime": "-",
        "contractNo": "-",
        "unitRaiseFunds": "-",
        "scienceFunding": "-"
    },
    "d": {
        "had_allocated": 0
    },
    "e": {
        "dishonest_money": 0
    },
    "f": {
        "no_allocated": 0
    },
    "aa": {
        "importedTechnology": "-",
        "applicationTechnology": "-",
        "scientificTechnologicalAchievementsTransformed": "-",
        "technicalTrading": "-",
        "newIndustrialProducts": "-",
        "newAgriculturalVariety": "-",
        "newProcess": "-",
        "newMaterial": "-",
        "newDevice": "-",
        "cs": "-",
        "researchPlatform": "-",
        "TS": "-",
        "pilotStudies": "-",
        "pilotLine": "-",
        "productionLine": "-",
        "experimentalBase": "-",
        "applyInventionPatent": "-",
        "applyUtilityModel": "-",
        "authorizedInventionPatent": "-",
        "authorizedUtilityModel": "-",
        "internationalStandard": "-",
        "nationalStandard": "-",
        "industryStandard": "-",
        "localStandards": "-",
        "enterpriseStandard": "-",
        "generalJournal": "-",
        "coreJournals": "-",
        "highLevelJournal": "-",
        "postdoctoralTraining": "-",
        "trainingDoctors": "-",
        "trainingMaster": "-",
        "monographs": "-",
        "academicReport": "-",
        "trainingCourses": "-",
        "trainingNumber": "-",
        "salesRevenue": "-",
        "newProduction": "-",
        "newTax": "-",
        "export": "-",
        "salesRevenue2": "-",
        "newProduction2": "-",
        "newTax2": "-",
        "export2": "-"
    }
},
{
    "annualPlan": "2021",
    "projectBatch": "第N批次",
    "planCategory": "科技重大专项",
    "projectName": "先进装备制造业1",
    "subjectName": "课题名称就是项目名称的名称",
    "unitName": "青秀区科技计划管理系统测试单位",
    "head": "刘一",
    "overallGoal": "课题总体目标就是为了测试",
    "startStopYear": "2021.06-2022.12",
    "assessmentIndicators": "1）完成第一个考核指标\n2）完成第二个考核指标\nx）完成第三个考核指标",
    "scienceFunding": 50000000,
    "unitSelfRaised": 0,
    "subjectState": "专家评审",
    "a": {
        "subjectScore": "-",
        "proposal": "-",
        "proposalFunding": "-"
    },
    "b": {
        "scienceProposal": "-",
        "scienceFunding": 0
    },
    "c": {
        "executionTime": "-",
        "contractNo": "-",
        "unitRaiseFunds": "-",
        "scienceFunding": "-"
    },
    "d": {
        "had_allocated": 0
    },
    "e": {
        "dishonest_money": 0
    },
    "f": {
        "no_allocated": 0
    },
    "aa": {
        "importedTechnology": "-",
        "applicationTechnology": "-",
        "scientificTechnologicalAchievementsTransformed": "-",
        "technicalTrading": "-",
        "newIndustrialProducts": "-",
        "newAgriculturalVariety": "-",
        "newProcess": "-",
        "newMaterial": "-",
        "newDevice": "-",
        "cs": "-",
        "researchPlatform": "-",
        "TS": "-",
        "pilotStudies": "-",
        "pilotLine": "-",
        "productionLine": "-",
        "experimentalBase": "-",
        "applyInventionPatent": "-",
        "applyUtilityModel": "-",
        "authorizedInventionPatent": "-",
        "authorizedUtilityModel": "-",
        "internationalStandard": "-",
        "nationalStandard": "-",
        "industryStandard": "-",
        "localStandards": "-",
        "enterpriseStandard": "-",
        "generalJournal": "-",
        "coreJournals": "-",
        "highLevelJournal": "-",
        "postdoctoralTraining": "-",
        "trainingDoctors": "-",
        "trainingMaster": "-",
        "monographs": "-",
        "academicReport": "-",
        "trainingCourses": "-",
        "trainingNumber": "-",
        "salesRevenue": "-",
        "newProduction": "-",
        "newTax": "-",
        "export": "-",
        "salesRevenue2": "-",
        "newProduction2": "-",
        "newTax2": "-",
        "export2": "-"
    }
}
]
def Excel_data(obj):
    # basedir = os.path.dirname(__file__)
    upload_path = MEDIA_ROOT + '数据导出.xlsx'
    wb = openpyxl.load_workbook(upload_path, data_only=True)
    sh = wb['Sheet1']
    print(sh.cell(1, 2).value)

    for i in range(len(obj)):
        x = 3
        for j in obj:
            print(j['subjectName'])
            sh.cell(x, 1).value = obj.index(j) + 1
            sh.cell(x, 2).value = j['annualPlan']
            sh.cell(x, 3).value = j['projectBatch']
            sh.cell(x, 4).value = j['planCategory']
            sh.cell(x, 5).value = j['projectName']
            sh.cell(x, 6).value = j['subjectName']
            sh.cell(x, 7).value = j['subjectState']
            sh.cell(x, 8).value = j['unitName']
            sh.cell(x, 9).value = j['head']
            sh.cell(x, 10).value = j['overallGoal']
            sh.cell(x, 11).value = j['startStopYear']
            sh.cell(x, 12).value = j['assessmentIndicators']
            sh.cell(x, 13).value = j['scienceFunding']
            sh.cell(x, 14).value = j['unitSelfRaised']

            sh.cell(x, 15).value = j['a']['subjectScore']
            sh.cell(x, 16).value = j['a']['proposal']
            sh.cell(x, 17).value = j['a']['proposalFunding']

            sh.cell(x, 18).value = j['b']['scienceProposal']
            sh.cell(x, 19).value = j['b']['scienceFunding']
            sh.cell(x, 20).value = j['c']['contractNo']
            sh.cell(x, 21).value = j['c']['executionTime']
            sh.cell(x, 22).value = j['c']['scienceFunding']
            sh.cell(x, 23).value = j['c']['unitRaiseFunds']

            sh.cell(x, 24).value = j['d']['had_allocated']

            sh.cell(x, 25).value = j['f']['no_allocated']

            sh.cell(x, 26).value = j['e']['dishonest_money']

            sh.cell(x, 27).value = j['aa']['importedTechnology']
            sh.cell(x, 28).value = j['aa']['applicationTechnology']
            sh.cell(x, 29).value = j['aa']['scientificTechnologicalAchievementsTransformed']
            sh.cell(x, 30).value = j['aa']['technicalTrading']
            sh.cell(x, 31).value = j['aa']['newIndustrialProducts']
            sh.cell(x, 32).value = j['aa']['newAgriculturalVariety']
            sh.cell(x, 33).value = j['aa']['newProcess']
            sh.cell(x, 34).value = j['aa']['newMaterial']
            sh.cell(x, 35).value = j['aa']['newDevice']
            sh.cell(x, 36).value = j['aa']['cs']
            sh.cell(x, 37).value = j['aa']['researchPlatform']
            sh.cell(x, 38).value = j['aa']['TS']
            sh.cell(x, 39).value = j['aa']['pilotStudies']
            sh.cell(x, 40).value = j['aa']['pilotLine']
            sh.cell(x, 41).value = j['aa']['productionLine']
            sh.cell(x, 42).value = j['aa']['experimentalBase']
            sh.cell(x, 43).value = j['aa']['applyInventionPatent']
            sh.cell(x, 44).value = j['aa']['applyUtilityModel']
            sh.cell(x, 45).value = j['aa']['authorizedInventionPatent']
            sh.cell(x, 46).value = j['aa']['authorizedUtilityModel']
            sh.cell(x, 47).value = j['aa']['internationalStandard']
            sh.cell(x, 48).value = j['aa']['nationalStandard']
            sh.cell(x, 49).value = j['aa']['industryStandard']
            sh.cell(x, 50).value = j['aa']['localStandards']
            sh.cell(x, 51).value = j['aa']['enterpriseStandard']
            sh.cell(x, 52).value = j['aa']['generalJournal']
            sh.cell(x, 53).value = j['aa']['coreJournals']
            sh.cell(x, 54).value = j['aa']['highLevelJournal']
            sh.cell(x, 55).value = j['aa']['postdoctoralTraining']
            sh.cell(x, 56).value = j['aa']['trainingDoctors']
            sh.cell(x, 57).value = j['aa']['trainingMaster']
            sh.cell(x, 58).value = j['aa']['monographs']
            sh.cell(x, 59).value = j['aa']['academicReport']
            sh.cell(x, 60).value = j['aa']['trainingCourses']
            sh.cell(x, 61).value = j['aa']['trainingNumber']
            sh.cell(x, 62).value = j['aa']['salesRevenue']
            sh.cell(x, 63).value = j['aa']['newProduction']
            sh.cell(x, 64).value = j['aa']['newTax']
            sh.cell(x, 65).value = j['aa']['export']
            sh.cell(x, 66).value = j['aa']['salesRevenue2']
            sh.cell(x, 67).value = j['aa']['newProduction2']
            sh.cell(x, 68).value = j['aa']['newTax2']
            sh.cell(x, 69).value = j['aa']['export2']
            x += 1
        wb.save(MEDIA_ROOT + 'a.xlsx')
        return MEDIA_ROOT + 'a.xlsx'






if __name__ == "__main__":
    print(Excel_data(a))